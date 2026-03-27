#!/usr/bin/env python3
"""
Resumen mensual de asistencia — Jibble → Excel + Discord
Royalspace 2026 | Port de jibble_resumen_mensual.ps1 xlsx1.4

Lógica replicada exactamente:
  - Mes anterior automáticamente
  - Excluye: Edwar, Angela Vanesa, Sebastian Reyes, Angelica Flores
  - Paginación por `time desc`, corte cuando time < inicio del mes
  - Primera marca "In" por persona por día
  - Scoring: A TIEMPO=0 | TARDE=-1.5 | FUERA DE RANGO=-2 | SIN MARCAR=-2
  - Fines de semana omitidos (columna vacía, no penaliza)
  - Empate UTC/local → prefiere UTC→Caracas (igual que el .ps1 mensual)
  - Excel con colores, freeze, negrita, bordes, fórmula SUM
  - Adjunta el .xlsx al mensaje de Discord
"""

from __future__ import annotations

import io
import json
import os
import sys
from calendar import monthrange
from datetime import datetime, timezone

import requests
import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from jibble_client import (
    VET,
    get_people,
    get_time_entries_page,
    get_token,
    parse_time_smart,
)

# ── Configuración ─────────────────────────────────────────────────────────────
CLIENT_ID       = os.environ["JIBBLE_CLIENT_ID"]
CLIENT_SECRET   = os.environ["JIBBLE_CLIENT_SECRET"]
ORG_ID          = os.environ["JIBBLE_ORG_ID"]
DISCORD_WEBHOOK = os.environ["DISCORD_WEBHOOK_ASISTENCIA"]

ON_TIME_END = (9, 15)
LATE_START  = (9, 16)
LATE_END    = (11, 0)

POINTS = {
    "ONTIME":   0.0,
    "LATE":    -1.5,
    "OUT":     -2.0,
    "MISSING": -2.0,
}

COUNT_WEEKENDS = False

EXCLUDE_NAMES = {"edwar", "angela vanesa", "sebastian reyes", "angelica flores"}

PAGE_SIZE = 200
MAX_PAGES = 300

MONTH_NAMES_ES = {
    1: "enero",    2: "febrero",  3: "marzo",     4: "abril",
    5: "mayo",     6: "junio",    7: "julio",      8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}

# Colores ARGB (formato openpyxl: AARRGGBB)
COLOR_GREEN       = "FFC6EFCE"   # 198,239,206
COLOR_YELLOW      = "FFFFF2CC"   # 255,242,204
COLOR_DARK_YELLOW = "FFFFD966"   # 255,217,102
COLOR_PASTEL_RED  = "FFF4CCCC"   # 244,204,204
COLOR_HEADER_BLUE = "FFB4C6E7"   # 180,198,231


# ── Helpers ───────────────────────────────────────────────────────────────────
def send_discord(title: str, body: str) -> None:
    requests.post(
        DISCORD_WEBHOOK,
        json={"content": f"**{title}**\n{body}"},
        timeout=15,
    ).raise_for_status()


def send_discord_with_file(title: str, body: str, file_bytes: bytes, filename: str) -> None:
    """Envía un mensaje de Discord con un archivo adjunto (.xlsx)."""
    requests.post(
        DISCORD_WEBHOOK,
        data={"payload_json": json.dumps({"content": f"**{title}**\n{body}"})},
        files={
            "file": (
                filename,
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        timeout=30,
    ).raise_for_status()


def classify(hour: int, minute: int) -> str:
    t = (hour, minute)
    if t <= ON_TIME_END:
        return "ONTIME"
    if LATE_START <= t <= LATE_END:
        return "LATE"
    return "OUT"


def is_workday(d: datetime) -> bool:
    """Lunes-Viernes = True. Fines de semana omitidos salvo COUNT_WEEKENDS=True."""
    return COUNT_WEEKENDS or d.weekday() < 5


# ── Excel ─────────────────────────────────────────────────────────────────────
def build_excel(
    people: list,
    states: dict[str, dict[int, str]],   # pid → {day_int → estado}
    year: int,
    month: int,
    days_in_month: int,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"

    # Congelar fila de encabezado
    ws.freeze_panes = "B2"

    thin    = Side(style="thin")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold    = Font(bold=True)

    fill_header      = PatternFill("solid", fgColor=COLOR_HEADER_BLUE)
    fill_map = {
        "ONTIME":  PatternFill("solid", fgColor=COLOR_GREEN),
        "LATE":    PatternFill("solid", fgColor=COLOR_YELLOW),
        "OUT":     PatternFill("solid", fgColor=COLOR_DARK_YELLOW),
        "MISSING": PatternFill("solid", fgColor=COLOR_PASTEL_RED),
    }

    # ── Fila 1: encabezados ──────────────────────────────────────────────────
    # Columna 1 = Nombre, columnas 2-32 = días 1-31, columna 33 = Total
    headers = ["Nombre"] + [str(d) for d in range(1, 32)] + ["Total"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = fill_header
        cell.font   = bold
        cell.border = border

    # ── Filas de datos ───────────────────────────────────────────────────────
    for row_idx, p in enumerate(people, start=2):
        pid  = str(p.get("id", ""))
        name = p.get("preferredName", "?")

        # Columna Nombre
        cell = ws.cell(row=row_idx, column=1, value=name)
        cell.border = border

        for d in range(1, 32):
            col = d + 1
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border

            if d > days_in_month:
                continue  # mes corto: dejar vacío

            if not is_workday(datetime(year, month, d)):
                continue  # fin de semana: vacío, no penaliza

            state = states.get(pid, {}).get(d)
            if state is None:
                continue

            cell.value = POINTS[state]
            cell.fill  = fill_map.get(state, PatternFill())

        # Columna Total (33): fórmula SUM sobre días 1-31 (columnas B a AF)
        total_cell = ws.cell(row=row_idx, column=33)
        total_cell.value  = f"=SUM(B{row_idx}:AF{row_idx})"
        total_cell.font   = bold
        total_cell.border = border

    # ── Auto-ajuste de columnas ──────────────────────────────────────────────
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[letter].width = max(max_len + 2, 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    now_caracas = datetime.now(timezone.utc).astimezone(VET)

    # Siempre mes anterior
    if now_caracas.month == 1:
        year, month = now_caracas.year - 1, 12
    else:
        year, month = now_caracas.year, now_caracas.month - 1

    days_in_month = monthrange(year, month)[1]

    # Límites del mes en VET (aware)
    month_start = VET.localize(datetime(year, month, 1, 0, 0, 0))
    next_month  = VET.localize(
        datetime(year if month < 12 else year + 1, month % 12 + 1, 1, 0, 0, 0)
    )
    cutoff_date = datetime(year, month, days_in_month)

    token = get_token(CLIENT_ID, CLIENT_SECRET)
    print("Token OK")

    people = [
        p for p in get_people(token, ORG_ID)
        if p.get("status") == "Joined"
        and (p.get("preferredName") or "").strip().lower() not in EXCLUDE_NAMES
    ]
    print(f"Personas activas: {len(people)}")

    # pid → {day_int → primera marca "In" del día}
    first_in: dict[str, dict[int, datetime]] = {str(p["id"]): {} for p in people}

    stop = False
    for page in range(MAX_PAGES):
        if stop:
            break

        items = get_time_entries_page(token, page * PAGE_SIZE, PAGE_SIZE, "time desc")
        if not items:
            break

        for e in items:
            if str(e.get("type", "")) != "In":
                continue

            # prefer_utc_fallback=True → empate: usar UTC→Caracas (igual que el .ps1 mensual)
            t = parse_time_smart(str(e.get("time", "")), prefer_utc_fallback=True)
            if not t:
                continue

            # Corte: cuando el tiempo ya es anterior al mes objetivo
            if t < month_start:
                stop = True
                break

            if t >= next_month:
                continue

            pid = str(e.get("personId", ""))
            if pid not in first_in:
                continue

            day = t.day
            if day not in first_in[pid] or t < first_in[pid][day]:
                first_in[pid][day] = t

    print("Entradas recolectadas")

    # ── Calcular estados por persona/día ──────────────────────────────────────
    states: dict[str, dict[int, str]] = {}
    for p in people:
        pid = str(p.get("id", ""))
        states[pid] = {}
        for d in range(1, days_in_month + 1):
            if not is_workday(datetime(year, month, d)):
                continue
            if d in first_in.get(pid, {}):
                cin   = first_in[pid][d]
                state = classify(cin.hour, cin.minute)
            else:
                state = "MISSING"
            states[pid][d] = state

    # ── Generar Excel ─────────────────────────────────────────────────────────
    month_name = MONTH_NAMES_ES[month]
    filename   = f"Asistencia_{year}-{month:02d}_{month_name}.xlsx"
    xlsx_bytes = build_excel(people, states, year, month, days_in_month)

    # Guardar en disco para que GitHub Actions lo suba como artifact
    out_path = os.path.join(os.path.dirname(__file__), filename)
    with open(out_path, "wb") as f:
        f.write(xlsx_bytes)
    print(f"Excel guardado: {out_path}")

    # ── Resumen para Discord (top 20 por total ascendente) ────────────────────
    scores: list[tuple[str, float]] = []
    for p in people:
        pid   = str(p.get("id", ""))
        name  = p.get("preferredName", "?")
        total = round(sum(POINTS.get(s, 0.0) for s in states.get(pid, {}).values()), 2)
        scores.append((name, total))

    scores.sort(key=lambda x: x[1])
    top20 = scores[:20]

    msg  = f"Mes: {month_name.capitalize()} {year}\n"
    msg += f"Corte: {cutoff_date.strftime('%d-%m-%Y')}\n\n"
    msg += "TOTAL:\n" + "\n".join(f"{name}: {total}" for name, total in top20)

    send_discord_with_file("[ASISTENCIA] Resumen mensual", msg, xlsx_bytes, filename)
    print("Mensaje + Excel enviados a Discord. OK")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        try:
            send_discord("[ASISTENCIA] ERROR (Resumen mensual)", f"Reporte mensual falló: {exc}")
        except Exception:
            pass
        sys.exit(1)
